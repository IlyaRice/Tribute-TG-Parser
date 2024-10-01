# parser.py
import re
import json
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import NamedStyle
import io

def process_data(file_obj):
    data = json.load(file_obj)

    result = []
    for message in data.get("messages", []):
        if message.get("from") != "Tribute":
            continue

        sender_name = extract_sender_name(message.get("text_entities", []))
        message_date = datetime.fromisoformat(message.get("date"))
        payment_type, amount = extract_payment_info(message.get("text_entities", []))

        if payment_type and amount is not None:
            result.append({
                "Дата и время": message_date,
                "Пользователь": sender_name,
                "Сумма": amount,
                "Категория": payment_type
            })

    df = create_dataframe_with_quarters(result)
    return df

def extract_sender_name(text_entities):
    for entity in text_entities:
        if entity.get("type") == "mention":
            return entity.get("text", "Неизвестно")
        elif entity.get("type") == "mention_name":
            name = entity.get("text", "Неизвестно")
            user_id = entity.get("user_id", "")
            return f"{name} (id{user_id})"
    return "Неизвестно"

def extract_payment_info(text_entities):
    payment_type = None
    amount = None

    full_text = ' '.join([entity.get("text", "").lower() for entity in text_entities])

    if "новая подписка" in full_text or "оформил подписку" in full_text:
        payment_type = "Новая подписка"
    elif "продлена подписка" in full_text or "продлил подписку" in full_text:
        payment_type = "Обновление подписки"
    elif "новый донат" in full_text or "отправил донат" in full_text or "отправил" in full_text:
        payment_type = "Донат"

    for entity in text_entities:
        if entity.get("type") == "bold":
            match = re.search(r"[₽]\d{1,7}\.\d{2}", entity.get("text", ""))
            if match:
                amount = float(match.group().replace('₽', '').replace('€', '').replace(',', '.'))
                break

    return payment_type, amount

def create_dataframe_with_quarters(data):
    df = pd.DataFrame(data)

    def get_quarter(month):
        if 1 <= month <= 3:
            return 'Q1'
        elif 4 <= month <= 6:
            return 'Q2'
        elif 7 <= month <= 9:
            return 'Q3'
        else:
            return 'Q4'

    df['Квартал'] = df['Дата и время'].apply(lambda x: get_quarter(x.month))
    return df

def style_excel_sheet(worksheet, start_row):
    # Apply styles starting from start_row (which includes the header row)
    text_format = '@'  # Text format
    date_style = NamedStyle(name='datetime', number_format='DD.MM.YYYY HH:MM')
    currency_style = NamedStyle(name='currency', number_format='# ##0.00 ₽')

    for row in worksheet.iter_rows(min_row=start_row + 1, max_col=4, max_row=worksheet.max_row):
        # Format 'Дата и время' column
        row[0].style = date_style

        # Format 'Пользователь' and 'Категория' columns as text
        for cell in [row[1], row[3]]:
            cell.number_format = text_format

        # Format 'Сумма' column
        row[2].style = currency_style

    # Set column widths and alignments
    worksheet.column_dimensions['A'].width = 18
    worksheet.column_dimensions['B'].width = 22
    worksheet.column_dimensions['C'].width = 14
    worksheet.column_dimensions['D'].width = 22

    # Align text to the left
    for col in worksheet.iter_cols(min_row=start_row, max_col=4, max_row=worksheet.max_row):
        for cell in col:
            cell.alignment = openpyxl.styles.Alignment(horizontal='left')


def calculate_summary(df):
    summary = {
        'Донат': [0, 0.0],
        'Новые и обновлённые подписки': [0, 0.0],
        'Налог 6% с подписок': [0, 0.0],
        'Всего платежей': [0, 0.0]
    }

    if 'Донат' in df['Категория'].values:
        donat_summary = df[df['Категория'] == 'Донат']['Сумма'].agg(['count', 'sum'])
        summary['Донат'] = [donat_summary['count'], donat_summary['sum']]

    subscription_df = df[df['Категория'].isin(['Новая подписка', 'Обновление подписки'])]
    if not subscription_df.empty:
        subscription_summary = subscription_df['Сумма'].agg(['count', 'sum'])
        summary['Новые и обновлённые подписки'] = [subscription_summary['count'], subscription_summary['sum']]
        summary['Налог 6% с подписок'] = [0, subscription_summary['sum'] * 0.06]

    summary['Всего платежей'] = [df.shape[0], df['Сумма'].sum()]

    return summary

def add_summary_to_excel(worksheet, summary):
    for r_idx, (key, values) in enumerate(summary.items(), 1):
        worksheet.cell(row=r_idx, column=1, value=key)

        for c_idx, value in enumerate(values, 2):
            cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = openpyxl.styles.Alignment(horizontal='left')
            if c_idx == 3:
                cell.number_format = '# ##0.00 ₽'

def save_to_excel(df):
    output_files = []
    quarter_summaries = {}
    quarter_to_months = {
        'Q1': '_янв_фев_мар',
        'Q2': '_апр_май_июн',
        'Q3': '_июл_авг_сен',
        'Q4': '_окт_ноя_дек'
    }

    for quarter in ['Q1', 'Q2', 'Q3', 'Q4']:
        quarter_df = df[df['Квартал'] == quarter]
        if quarter_df.empty:
            continue

        quarter_df = quarter_df.drop(columns=['Квартал'])
        summary = calculate_summary(quarter_df)

        # Create a DataFrame for the summary
        summary_df = pd.DataFrame({
            'Категория': summary.keys(),
            'Количество': [v[0] for v in summary.values()],
            'Сумма': [v[1] for v in summary.values()]
        })

        formatted_summary = '\n'.join([
            f"{key} - {int(value[0])}шт, {value[1]:,.2f}₽" if key != 'Налог 6% с подписок' else f"{key} - {value[1]:,.2f}₽"
            for key, value in summary.items()
        ])
        quarter_summaries[quarter + quarter_to_months[quarter]] = formatted_summary

        output = io.BytesIO()
        sheet_name = f'Квартал {quarter}{quarter_to_months[quarter]}'
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write the summary DataFrame to Excel
            summary_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)

            # Write an empty row between summary and data
            startrow = summary_df.shape[0] + 2  # Leave one empty row
            # Write the quarter_df DataFrame starting from startrow
            quarter_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)

            # Access the worksheet to apply formatting
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Apply styles to the data
            style_excel_sheet(worksheet, start_row=startrow + 1)  # +1 to account for the header row

            # Format 'Сумма' column in the summary
            for row in range(2, summary_df.shape[0] + 2):  # Rows start at 2 due to header
                cell = worksheet.cell(row=row, column=3)  # 'Сумма' is the third column
                cell.number_format = '# ##0.00 ₽'

        output.seek(0)
        output_files.append((f'Квартал_{quarter}{quarter_to_months[quarter]}.xlsx', output))

    return output_files, quarter_summaries